from selenium.webdriver.common.by import By
from tpffpsldna.Config.Accounts import *
from tpffpsldna.Pages.Base import BaseItem
from openpyxl import *

import time

class JD(BaseItem):

    text = (By.CSS_SELECTOR,"#content > div > div > div.cnt-list-wrap > div > div.recruit-info > div.list-filter-wrap > p")
    wait = (By.CSS_SELECTOR,".post-list-corp")

    def __init__(self, driver):
        super(JD,self).__init__(driver)
        self.driver = self.inter.driver

    def JD_Crawling(self):
        # 채용 공고 전체 수 파악
        JD_total = self.driver.find_element(*self.text).text
        if JD_total:
            print("검색된 채용 공고는" + JD_total + "입니다.")
        else:
            raise Exception("[ERROR] 현재 노출되지 않습니다.")
        
        #---- 채용 공고 크롤링
        m = range(2,3)

        for s in m:
            

            a = self.inter.driver.find_element(By.CSS_SELECTOR,".list-default").find_elements(By.CSS_SELECTOR,".list-post")
            
            Compuny = []
            title = []
            local = []
            day = []
            
            total = []

            for i in (a):
                
                # 기업명
                list1 = i.find_element(By.CSS_SELECTOR,".name.dev_view").text
                Compuny.append(list1)
                
                # 채용 공고 제목
                list2 = i.find_element(By.CSS_SELECTOR,".title.dev_view").text
                title.append(list2)

                # 근무 지역
                list3 = i.find_element(By.CSS_SELECTOR,".loc.long").text
                local.append(list3)

                # 공고 기한
                list4 = i.find_element(By.CSS_SELECTOR,".date").text
                day.append(list4)

                # 전체 크롤링 리스트
                total.append([list1, list2, list3, list4])
            
            time.sleep(4)
            # 다음 페이지 
            self.inter.driver.find_element(By.CSS_SELECTOR,"[page-no = '%d']"%s).click()
            time.sleep(4)
           
            # wb = Workbook()
            
            # 엑셀 파일 불러오기
            wb = load_workbook('./ex.xlsx')
            ws = wb.active
           
            # 시트 생성
            # wb.create_sheet(index=3,title=sheet_title)
           
            
            # 1행 1열 , 2열... 열에 문자열 삽입
            ws.cell(row=1, column=1,value = "기업명")
            ws.cell(row=1, column=2,value = "채용 공고 제목")
            
            for data in total: # 엑셀 저장 1줄씩 생성
                ws.append(data)

            wb.save('./ex.xlsx')
    

